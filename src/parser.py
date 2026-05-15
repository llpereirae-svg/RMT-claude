"""Parser del Resumen Mensual de Transacciones exportado a Excel."""
from __future__ import annotations

import hashlib
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from .models import AnuladoRMT, DocumentoRMT, RMTProcesado, RetencionRMT

SECTION_NAMES = {
    "FACTURAS EMITIDAS:",
    "NOTAS DE CREDITO EMITIDAS:",
    "RETENCIONES EMITIDAS:",
    "NOTAS DE DEBITO EMITIDAS:",
    "DOCUMENTOS RECIBIDOS ATS:",
    "FACTURAS RECIBIDAS (CEDULA):",
    "DOCUMENTOS RECIBIDOS QUE NO VAN EN EL ATS:",
    "GASTOS DE VIAJE:",
    "LIQUIDACION DE COMPRA:",
    "NOTAS DE CREDITO RECIBIDAS:",
    "NOTAS DE DEBITO RECIBIDAS:",
    "RETENCIONES RECIBIDAS:",
    "LIQUIDACION ADUANERA:",
}

HEADER_ROWS = {
    "FACTURAS EMITIDAS:": 1,
    "NOTAS DE CREDITO EMITIDAS:": 1,
    "RETENCIONES EMITIDAS:": 1,
    "NOTAS DE DEBITO EMITIDAS:": 1,
    "DOCUMENTOS RECIBIDOS ATS:": 1,
    "FACTURAS RECIBIDAS (CEDULA):": 1,
    "DOCUMENTOS RECIBIDOS QUE NO VAN EN EL ATS:": 1,
    "GASTOS DE VIAJE:": 1,
    "LIQUIDACION DE COMPRA:": 1,
    "NOTAS DE CREDITO RECIBIDAS:": 1,
    "NOTAS DE DEBITO RECIBIDAS:": 1,
    "RETENCIONES RECIBIDAS:": 1,
    "LIQUIDACION ADUANERA:": 1,
}


def hash_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def parse(path: Path) -> RMTProcesado:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Por ahora el parser implementado soporta RMT en Excel (.xlsx/.xlsm).")

    # El RMT es pequeño/mediano, pero read_only penaliza mucho el acceso por celda
    # en bloques. Cargar normal mantiene el parser por debajo de segundos.
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb.active
    sections = _find_sections(ws)
    ruc, nombre = _read_company(ws)
    periodo_inicio, periodo_fin = _read_periods(ws)
    parsed = RMTProcesado(
        archivo=path,
        nombre_archivo=path.name,
        hash_archivo=hash_file(path),
        ruc=ruc,
        cliente_nombre=nombre,
        periodo_inicio=periodo_inicio,
        periodo_fin=periodo_fin,
    )

    for idx, (section, row) in enumerate(sections):
        next_row = sections[idx + 1][1] if idx + 1 < len(sections) else ws.max_row + 1
        header_row = row + HEADER_ROWS.get(section, 1)
        headers = _headers(ws, header_row)
        for excel_row in range(header_row + 1, next_row):
            values = _row_dict(ws, excel_row, headers)
            if _skip_row(values):
                continue
            try:
                _append_row(parsed, section, values)
            except Exception as exc:  # pragma: no cover - defensive branch for malformed RMT files
                parsed.advertencias.append(f"{section} fila {excel_row}: no se pudo leer ({exc}).")

    _post_validate(parsed)
    return parsed


def _find_sections(ws: Any) -> list[tuple[str, int]]:
    found: list[tuple[str, int]] = []
    for row in range(1, ws.max_row + 1):
        first = _text(ws.cell(row, 1).value).upper()
        row_text = " ".join(_text(ws.cell(row, col).value).upper() for col in range(1, min(ws.max_column, 5) + 1))
        for name in SECTION_NAMES:
            if first == name or name in row_text:
                found.append((name, row))
                break
    return found


def _read_company(ws: Any) -> tuple[str, str]:
    for row in range(1, min(ws.max_row, 10) + 1):
        if _text(ws.cell(row, 1).value).upper() == "EMPRESA":
            ruc = _digits(ws.cell(row, 2).value)
            if len(ruc) == 10:
                ruc = f"{ruc}001"
            nombre = _text(ws.cell(row, 3).value)
            return ruc, nombre
    return "", ""


def _read_periods(ws: Any) -> tuple[str, str]:
    for row in range(1, min(ws.max_row, 8) + 1):
        if _text(ws.cell(row, 1).value).upper() == "FECHA INI":
            raw = "".join(_text(ws.cell(row, col).value) for col in range(2, min(ws.max_column, 5) + 1))
            matches = re.findall(r"20\d{2}/\d{1,2}", raw)
            if len(matches) >= 2:
                return _period(matches[0]), _period(matches[1])
            if len(matches) == 1:
                return _period(matches[0]), _period(matches[0])
    return "", ""


def _headers(ws: Any, row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    seen: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        label = _text(ws.cell(row, col).value).upper()
        if not label:
            continue
        seen[label] = seen.get(label, 0) + 1
        key = label if seen[label] == 1 else f"{label}#{seen[label]}"
        headers[key] = col
    return headers


def _row_dict(ws: Any, row: int, headers: dict[str, int]) -> dict[str, Any]:
    return {name: ws.cell(row, col).value for name, col in headers.items()}


def _append_row(parsed: RMTProcesado, section: str, row: dict[str, Any]) -> None:
    if section in {"RETENCIONES EMITIDAS:", "RETENCIONES RECIBIDAS:"}:
        parsed.retenciones.append(_retencion(section, row))
        return
    if section == "GASTOS DE VIAJE:":
        parsed.documentos.append(_gasto_viaje(row))
        return
    if section == "LIQUIDACION ADUANERA:":
        parsed.documentos.append(_aduana(row))
        return
    if _is_annulled(row):
        parsed.anulados.append(_anulado(section, row))
        return
    parsed.documentos.append(_documento(section, row))


def _documento(section: str, row: dict[str, Any]) -> DocumentoRMT:
    numero = _first(row, "FACTURA", "NOTA DE CREDITO", "NOTA DE DEBITO", "NUM-DOC")
    contraparte_id = _clean_party(_first(row, "CLIENTE CI/RUC", "PROVEEDOR CI/RUC"))
    contraparte_nombre = _text(_first(row, "CLIENTE RAZON SOCIAL", "PROVEEDOR RAZON SOCIAL"))
    return DocumentoRMT(
        bloque=section.replace(":", ""),
        grupo=_classify_group(section),
        numero=_text(numero),
        fecha_emision=_date(_first(row, "FECHA EMISION")),
        fecha_carga=_datetime(_first(row, "FECHA DE CARGA")),
        contraparte_id=contraparte_id,
        contraparte_nombre=contraparte_nombre,
        base_0=_num(_first(row, "BASE 0%", "BASE 0")),
        descuento_0=_num(_first(row, "DESCUENTO 0%")),
        base_iva=_num(_first(row, "BASE 12%", "BASE IVA%", "BASE IVA")),
        descuento_iva=_num(_first(row, "DESCUENTO 12%")),
        no_objeto_iva=_num(_first(row, "NO OBJETO DE IVA")),
        descuento_no_objeto=_num(_first(row, "DESCUENTO NO OBJETO DE IVA")),
        total_sin_impuesto=_num(_first(row, "TOTAL SIN IMPUESTO")),
        iva=_num(_first(row, "% IVA", "IVA 12%", "IVA TOTAL%", "IVA GENER", "IVA PAGADO")),
        servicio=_num(_first(row, "% SERVICIO", "OTROS/PROPINA")),
        propina=_num(_first(row, "PROPINA")),
        total=_num(_first(row, "TOTAL")),
        sustento=_text(_first(row, "SUSTENTO CRED. TRIB.", "R.GASTOS/EXPORT.")),
        autorizacion=_clean_authorization(_first(row, "NUMERO AUTORIZACION", "AUTORIZACION")),
        doc_modificado=_clean_doc(_first(row, "NUMERO DE DOC MODIFICADO")),
        cuadre=_num(_first(row, "CUADRE")),
        extra={k: _text(v) for k, v in row.items() if v not in (None, "")},
    )


def _gasto_viaje(row: dict[str, Any]) -> DocumentoRMT:
    return DocumentoRMT(
        bloque="GASTOS DE VIAJE",
        grupo="compra",
        numero=f"{_text(_first(row, 'SERIE'))}-{_text(_first(row, 'FACTURA'))}".strip("-"),
        fecha_emision=_date(_first(row, "FECHA")),
        fecha_carga=_datetime(_first(row, "FECHA")),
        contraparte_id=_clean_party(_first(row, "RUC")),
        contraparte_nombre=_text(_first(row, "RAZON SOCIAL")),
        base_0=_num(_first(row, "BASE 0")),
        base_iva=_num(_first(row, "BASE IVA")),
        iva=_num(_first(row, "IVA")),
        propina=_num(_first(row, "OTROS/PROPINA")),
        total=_num(_first(row, "TOTAL")),
        sustento=_text(_first(row, "GASTO")),
        autorizacion=_clean_authorization(_first(row, "AUTORIZACION")),
        extra={k: _text(v) for k, v in row.items() if v not in (None, "")},
    )


def _aduana(row: dict[str, Any]) -> DocumentoRMT:
    return DocumentoRMT(
        bloque="LIQUIDACION ADUANERA",
        grupo="importacion",
        numero=_text(_first(row, "NUM-DOC", "NUM REGISTRO")),
        fecha_emision=_date(_first(row, "FECHA EMISION")),
        fecha_carga=_datetime(_first(row, "FECHA DE CARGA")),
        contraparte_id=_clean_party(_first(row, "PROVEEDOR CI/RUC")),
        contraparte_nombre=_text(_first(row, "PROVEEDOR RAZON SOCIAL")),
        base_0=_num(_first(row, "BASE 0%")),
        base_iva=_num(_first(row, "BASE IVA%", "BASES")),
        iva=_num(_first(row, "IVA PAGADO", "IVA GENER")),
        total=_num(_first(row, "T.PAGO")),
        cuadre=_num(_first(row, "CUADRE")),
        extra={k: _text(v) for k, v in row.items() if v not in (None, "")},
    )


def _retencion(section: str, row: dict[str, Any]) -> RetencionRMT:
    codigos = [v for k, v in row.items() if k.startswith("CODIGO") and _text(v)]
    bases = [v for k, v in row.items() if k.startswith("BASE IMPONIBLE") and _text(v)]
    valores = [v for k, v in row.items() if k.startswith("VALOR RETENCION") and _text(v)]
    return RetencionRMT(
        bloque=section.replace(":", ""),
        numero=_text(_first(row, "RETENCION")),
        fecha_emision=_date(_first(row, "FECHA EMISION")),
        contraparte_id=_clean_party(_first(row, "PROVEEDOR CI/RUC")),
        contraparte_nombre=_text(_first(row, "PROVEEDOR RAZON SOCIAL")),
        factura=_text(_first(row, "No. FACTURA PROVEEDOR", "No. FACTURA")),
        iva_10=_num(_first(row, "IVA 10%")),
        iva_20=_num(_first(row, "IVA 20%")),
        iva_30=_num(_first(row, "IVA 30%")),
        iva_50=_num(_first(row, "IVA 50%")),
        iva_70=_num(_first(row, "IVA 70%")),
        iva_100=_num(_first(row, "IVA 100%")),
        codigo_ir=_text(codigos[0]) if codigos else "",
        base_ir=_num(bases[0]) if bases else 0.0,
        valor_ir=_num(valores[0]) if valores else 0.0,
        extra={k: _text(v) for k, v in row.items() if v not in (None, "")},
    )


def _anulado(section: str, row: dict[str, Any]) -> AnuladoRMT:
    return AnuladoRMT(
        tipo=section.replace(":", ""),
        fecha=_date(_first(row, "FECHA EMISION")),
        numero=_text(_first(row, "FACTURA", "NOTA DE CREDITO", "NOTA DE DEBITO", "NUM-DOC")),
        autorizacion=_clean_authorization(_first(row, "NUMERO AUTORIZACION", "AUTORIZACION")),
        contraparte_id=_clean_party(_first(row, "CLIENTE CI/RUC", "PROVEEDOR CI/RUC")),
        contraparte_nombre=_text(_first(row, "CLIENTE RAZON SOCIAL", "PROVEEDOR RAZON SOCIAL")),
        total=_num(_first(row, "TOTAL")),
    )


def _post_validate(parsed: RMTProcesado) -> None:
    for doc in parsed.documentos:
        if doc.bloque == "LIQUIDACION ADUANERA" and abs(doc.cuadre) > 20:
            parsed.advertencias.append(
                f"Liquidacion aduanera {doc.numero}: CUADRE {doc.cuadre:.2f} supera +/-20."
            )
        if doc.sustento.strip().upper() == "E":
            parsed.advertencias.append(
                f"Exportacion detectada en {doc.numero}: revisar ATS para clasificar bienes/servicios."
            )


def _skip_row(row: dict[str, Any]) -> bool:
    values = [_text(v) for v in row.values() if _text(v)]
    if not values:
        return True
    first = values[0].lower()
    return first in {"totales", "total"} or len(values) == 1


def _is_annulled(row: dict[str, Any]) -> bool:
    return any(_text(v).upper() == "ANULADO" for v in row.values())


def _classify_group(section: str) -> str:
    if section in {"FACTURAS EMITIDAS:", "NOTAS DE DEBITO EMITIDAS:"}:
        return "venta"
    if section == "NOTAS DE CREDITO EMITIDAS:":
        return "nc_venta"
    if section == "NOTAS DE CREDITO RECIBIDAS:":
        return "nc_compra"
    if section == "NOTAS DE DEBITO RECIBIDAS:":
        return "compra"
    if section == "LIQUIDACION ADUANERA:":
        return "importacion"
    return "compra"


def _first(row: dict[str, Any], *names: str) -> Any:
    for name in names:
        if name in row:
            return row[name]
    return None


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _num(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = _text(value).replace(",", "")
    try:
        return round(float(text), 2)
    except ValueError:
        return 0.0


def _digits(value: Any) -> str:
    return "".join(re.findall(r"\d+", _text(value)))


def _clean_party(value: Any) -> str:
    text = _text(value)
    if text.upper().startswith("N "):
        text = text[2:]
    return _digits(text)


def _clean_authorization(value: Any) -> str:
    text = _text(value)
    if text.upper().startswith("A "):
        return text[2:].strip()
    return text


def _clean_doc(value: Any) -> str:
    text = _text(value)
    if text.upper().startswith("N "):
        return text[2:].strip()
    return text


def _date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S%z"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    if " " in text:
        return _date(text.split(" ")[0])
    return None


def _datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = _text(value)
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S%z", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _period(value: str) -> str:
    year, month = value.split("/")
    return f"{int(year):04d}-{int(month):02d}"
